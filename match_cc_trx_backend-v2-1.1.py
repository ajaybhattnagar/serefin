
from datetime import datetime
import linecache
import configparser
import pyodbc
import os
import sys
import platform
import getopt
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

import serefin

#
# This process will update the WRK_HSV_CC_TRANSACTION table with information identified in the PenAir backend.
#
# 1.  Workbook to be generated daily
# 2.  New TAB added to the existing monthly workbook to reprecent the previoues days activities
# 3.  Report is to be mailed to the stake holders

# NOTE:  DO NOTE PERFORM A REFORMAT OF FILE WITHIN PYCHARM AS IT WILL CHANGE THE CASE TESTING AND EFFECT
#        THE RESULTS OF THE REPORT (EFFECTS PEP   -1 )
#
# Modifications
# 2023-09-06  FSB  1: Initial creation
# 2023-09-18  FSB  1: Added product type to transaction details line
# 2023-09-21  FSB  1: Added agent name to transaction line
#                  2: Add supplier name to transaction details line
#                  3: Add Web Reference to the transaction details line
# 2023-10-19  FSB 1: Added partial lookup with travel date and last PAX name
# 2023-11-06  FSB 1: Changed select SQL from using the Date_Occurred to the Date_Posted
# 2023-11-15  FSB 1: Remove date selection range from base processing
#                 2: Remove the restriction of the no PAX name and ticket number
#                 3: Add extra processing based on the merchant name like TRAVEL 9012344567
#                    as the 2nd string is the Your Ref
# 2024-01-03  FSB 1: Removed FolderMaster Status restriction on all select statements
#                 2: Removed FolderMaster BAID on selects where full project no and folder no are part of parameters
# 2024-01-05  FSB 1: Added automatic workbook generation and email process
# 2024-01-11  FSB 1: Tweak SQL
# 2024-01-30  FSB 1: Changed code to use SQL Server instead of MySQL.
#                  : Chnages mage are to replace the %s to ?, now() to CURRENT_TIMESTAMP, LENGTH() to LEN()
# 2024-04-18  FSB 1: Added handling of PREFERRED ZONE transactions to identify lead's
#

#
# Load Configuration Information from INI file

config = configparser.ConfigParser()
config.sections()

if platform.release() == '10':
    config.read('/dev/testing_config.ini')
else:
    config.read(os.path.basename(sys.argv[0]).replace('.py', '.ini'))
        
#
# handle the MS SQL Server Information
msserver = config['BIConfig']
ms_svr = msserver['ServerIP']
ms_db = msserver['DBName']
ms_db_na = msserver['DBNameNA']
ms_db_drv = msserver['DBDriver']
ms_usr = msserver['DBUser']
ms_pwd = msserver['DBUserPwd']

#
# handle the MS SQL Server Information
msserver = config['MSConfig']
msp_svr = msserver['ServerIP']
msp_db_na = msserver['DBNameNA']
msp_db_drv = msserver['DBDriver']
msp_usr = msserver['DBUser']
msp_pwd = msserver['DBUserPwd']

# Handle Email Server information
emailserver = config['EmailServer']
email_server = emailserver['Server']
email_user = emailserver['User']
email_from = emailserver['Sender']
email_pass = emailserver['UserPass']

# Handle Destination Email Details
emaildestination = config['EmailDestination']
email_to = emaildestination['Destination']
email_cc = emaildestination['CarbineCopy']

#
# Main code
def print_exception():
    """This method is use to trace and print exception"""
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    line_no = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, line_no, f.f_globals)
    print(f"EXCEPTION IN ({filename}, LINE {line_no} '{line.strip()}'):\n {exc_obj}")


def buildheadings(worksheet):
    """Create Formatting  Header and Data Formatting"""
    # Transaction_ID, Record_Type, Merchant, Merchant_Location, Account_Number,
    # MCC, Date_Occurred, Date_Posted, Original_Amount, Original_Currency_Code,
    # Conversion_Rate, Settlement_Amount, Allocation, Transaction_Description, Reference_Number,
    # Purch_Description, Purch_Quantity, Purch_Unit_Cost, Purch_Unit_Of_Measure, Purch_Extended_Amount,
    # Passenger_Name, Ticket_Number, Travel_Date, Travel_Legs, import_date,
    # projno as folderno, folderno as orderid, leadid, contract_number, product_type,
    # webref, booking_agent, supplier
    #
    #  Definition per column : [Row, Column No, Title, Column Letter, length]
    #
    # Create the work sheet header columns
    headers = [[1, 1, 'Transaction Id', 'A', 15], [1, 2, 'Record Type', 'B', 15], [1, 3, 'Merchant', 'C', 30],
        [1, 4, 'Merchant Location', 'D', 30], [1, 5, 'Account Number', 'E', 20], [1, 6, 'MMC', 'F', 5],
        [1, 7, 'Date Occurred', 'G', 15], [1, 8, 'Date Posted', 'H', 15], [1, 9, 'Original Amount', 'I', 18],
        [1, 10, 'Original Currency Code', 'J', 10], [1, 11, 'Conversion_Rate', 'K', 15],
        [1, 12, 'Settlement_Amount', 'L', 15], [1, 13, 'Allocation', 'M', 15],
        [1, 14, 'Transaction_Description', 'N', 10], [1, 15, 'Reference_Number', 'O', 20],
        [1, 16, 'Purch_Description', 'P', 15], [1, 17, 'Purch_Quantity', 'Q', 15],
        [1, 18, 'Purch_Unit_Cost', 'R', 15], [1, 19, 'Purch_Unit_Of_Measure', 'S', 15],
        [1, 20, 'Purch_Extended_Amount', 'T', 15], [1, 21, 'Passenger_Name', 'U', 30],
        [1, 22, 'Ticket_Number', 'V', 20], [1, 23, 'Travel_Date', 'W', 15], [1, 24, 'Travel_Legs', 'X', 15],
        [1, 25, 'import_date', 'Y', 15], [1, 26, 'Folder No', 'Z', 15], [1, 27, 'Order No', 'AA', 15],
        [1, 28, 'Lead Id', 'AB', 15], [1, 29, 'Contract Number', 'AC', 15], [1, 30, 'Product Type', 'AD', 15],
        [1, 31, 'Web Ref', 'AE', 15], [1, 32, 'Booking Agent', 'AF', 15], [1, 33, 'Supplier', 'AG', 15]]

    for head in headers:
        worksheet.cell(row=head[0], column=head[1]).value = head[2]
        worksheet.column_dimensions[head[3]].width = head[4]

    # Change the colours of the headings
    for c in worksheet['A1:AG1'][0]:
        c.fill = PatternFill('solid', start_color='009999FF')

    for c in worksheet['A1:AG1'][0]:
        c.alignment = Alignment(horizontal='center', vertical='center')

    #
    # worksheet.freeze_panes = 'Q1'
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    return


def getLeadMemberShipId(key_value, cur_worker):
    """Collect the HSV lead and Membership fields from the CS fields table"""

    sel_sql = """select fm.projno, 
    replace(replace(replace(replace(n0.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as leadid,
    replace(replace(replace(replace(n1.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as membernum, 
    n2.fcsf_labeldata as contractnum, n3.fcsf_labeldata as source
    from foldermaster as fm
left join FolderCustomerSpecificField_Table as n0 on n0.fcsf_projno = fm.projno
            and n0.fcsf_folderno = fm.folderno and n0.fcsf_labelname = 'LeadID'
left join FolderCustomerSpecificField_Table as n1 on n1.fcsf_projno = fm.projno
            and n1.fcsf_folderno = fm.folderno and n1.fcsf_labelname = 'Membership Number'
left join FolderCustomerSpecificField_Table as n2 on n2.fcsf_projno = fm.projno
           and n2.fcsf_folderno = fm.folderno and n2.fcsf_labelname = 'ContractNumber'
left join FolderCustomerSpecificField_Table as n3 on n3.fcsf_projno = fm.projno
      and n3.fcsf_folderno = fm.folderno and n3.fcsf_labelname = 'Source'
where fm.projno = '{0}' and fm.folderno = '{1}'    
    """
    cur_worker.execute(sel_sql.format(key_value[0], key_value[1]))
    row = cur_worker.fetchone()
    # 0 PROJNO
    # 1 lead id
    # 2 member number
    # 3 contract number
    # 4 source
    # 5 DDH lead Id
    # 6 DDH Member number
    the_leadid = ''
    the_contract = ''
    if row:
        if row[1] is None or row[1] == '':
            if row[2] is None or row[2] == '':
                the_leadid = ''
            else:
                the_leadid = row[2]
        else:
            the_leadid = row[1]
        if row[3]:
            the_contract = row[3]
        else:
            if row[4]:
                if 'US' in row[4]:
                    the_contract = row[4][2:]
                else:
                    the_contract = row[4]

    results = [the_leadid, the_contract]
    return results


def getContractNumber(key_value, cur_worker):
    """Check the HSV supplied table for the contract number"""

    sel_sql = f"select contract_number from wrk_hsv_lead_contract where leadid = '{key_value}' ;"
    cur_worker.execute(sel_sql)
    row = cur_worker.fetchone()

    if row is None:
        results = ''
    else:
        results = row[0]
    return results


def getLeadid(key_value, cur_worker):
    """Check the HSV supplied table for the lead ID using the contract number"""

    sel_sql = f"select leadid from wrk_hsv_lead_contract where contract_number = '{key_value}' ;"
    cur_worker.execute(sel_sql)
    row = cur_worker.fetchone()

    if row is None:
        results = ''
    else:
        results = row[0]
    return results


def getFolderMasterDetails(key_value, cur_worker):
    """Get any data from the FolderMaster table based on the YourRef/PONO attribute"""

    sel_sql = """select fm.projno, fm.folderno, fm.inetref, fm.pass, fm.bookedby, fm.folderdate from FolderMaster as fm 
    where fm.pono = '{0}'; """

    cur_worker.execute(sel_sql.format(key_value))
    # row = cur_worker.fetchall()

    # cur_worker.execute(sel_sql.format(key_value))
    row = cur_worker.fetchone()
    results = ''

    if row is None:
        results = ''
    else:
        results = row
    return results


def getFolderMasterbyProjno(key_value, cur_worker):
    """Get any data from the FolderMaster table based on the YourRef/PONO attribute"""

    sel_sql = """select fm.projno, fm.folderno, fm.inetref, fm.pass, fm.bookedby, fm.folderdate from FolderMaster as fm 
    where fm.projno = '{0}'; """

    cur_worker.execute(sel_sql.format(key_value))
    # row = cur_worker.fetchall()

    # cur_worker.execute(sel_sql.format(key_value))
    row = cur_worker.fetchone()
    results = ''

    if row is None:
        results = ''
    else:
        results = row
    return results


def getOtherdetails(key_value, cur_worker):
    """Get any data from the OthersDetail table that was likely charged and may appear on report"""

    sel_sql = """select od.projno, od.folderno,fm.inetref, od.suppname, od.description, od.productdescription, 
    od.bookedby, od.fcbuy, od.hcsell as Sell, od.hccommamt as commamt, convert(date, od.bookingdate) as BookingDate,
    od.loyaltypointsmembno
from OthersDetails as od   
inner join foldermaster as fm on od.projno = fm.projno and od.folderno = fm.folderno         
where od.projno = '{0}' and od.folderno = '{1}' """

    cur_worker.execute(sel_sql.format(key_value[0], key_value[1]))
    row = cur_worker.fetchone()
    results = ''

    if row is None:
        results = ''
    else:
        results = row
    return results


def getLeadContractDetails(key_values, cur_worker):
    """Collect the ticket details based on the passenger last name"""

    sel_sql = """
        select top 1 replace(replace(replace(replace(n1.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as LeadId1,
           n2.fcsf_labeldata as ContractNum, n3.fcsf_labeldata as Source,
           replace(replace(replace(replace(n4.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as LeadId2
    FROM FolderCustomerSpecificField_Table as t0
    inner join foldermaster as fm on fm.projno = t0.fcsf_projno and fm.folderno = t0.fcsf_folderno
    left join FolderCustomerSpecificField_Table as n1 on n1.fcsf_projno = fm.projno
          and n1.fcsf_folderno = fm.folderno and n1.fcsf_labelname = 'Membership Number'
    left join FolderCustomerSpecificField_Table as n2 on n2.fcsf_projno = fm.projno
          and n2.fcsf_folderno = fm.folderno and n2.fcsf_labelname = 'ContractNumber'
    left join FolderCustomerSpecificField_Table as n3 on n3.fcsf_projno = fm.projno
          and n3.fcsf_folderno = fm.folderno and n3.fcsf_labelname = 'Source'
    left join FolderCustomerSpecificField_Table as n4 on n4.fcsf_projno = fm.projno
          and n4.fcsf_folderno = fm.folderno and n4.fcsf_labelname = 'DDH Lead Id'
    WHERE t0.fcsf_projno = '{0}' and t0.fcsf_folderno = '{1}'; """

    #  LeadId1, ContractNumber, Source, LeadId2
    #      0           1           2       3
    cur_worker.execute(sel_sql.format(key_values[0], key_values[1]))
    row = cur_worker.fetchone()

    the_leadid = ''
    the_contract = ''
    results = ''

    while row:
        # 4  = Lead Id
        # 5  = ContractNum
        # 6  = Source
        # 7  = LeadId4,
        # 8  = DDHMemberNum
        if row[0] is None or row[0] == '':
            if row[3] is None or row[3] == '':
                the_leadid = ''
            else:
                the_leadid = row[3]
        else:
            the_leadid = row[0]

        if row[1]:
            the_contract = row[1]
        else:
            if row[2]:
                if 'US' in row[2]:
                    the_contract = row[2][2:]
                else:
                    the_contract = row[2]

        # if the_leadid
        results = [the_leadid, the_contract]
        row = cur_worker.fetchone()
    return results


def getTravelDate(key_value, cur_worker):
    """Collect the ticket details based on the travel date and the passenger possible name"""

    sel_sql = """
    select atd.projno, atd.folderno, atd.passname, atd.ticketno,
       replace(replace(replace(replace(n1.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as LeadId,
       n2.fcsf_labeldata as ContractNum, n3.fcsf_labeldata as Source,
       replace(replace(replace(replace(n4.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as LeadId4,
       n5.fcsf_labeldata as DDHMemberNum
from AirTicketDetails as atd
inner join foldermaster as fm on fm.projno=atd.projno and fm.folderno = atd.folderno
left join FolderCustomerSpecificField_Table as n1 on n1.fcsf_projno = fm.projno
      and n1.fcsf_folderno = fm.folderno and n1.fcsf_labelname = 'Membership Number'
left join FolderCustomerSpecificField_Table as n2 on n2.fcsf_projno = fm.projno
      and n2.fcsf_folderno = fm.folderno and n2.fcsf_labelname = 'ContractNumber'
left join FolderCustomerSpecificField_Table as n3 on n3.fcsf_projno = fm.projno
      and n3.fcsf_folderno = fm.folderno and n3.fcsf_labelname = 'Source'
left join FolderCustomerSpecificField_Table as n4 on n4.fcsf_projno = fm.projno
      and n4.fcsf_folderno = fm.folderno and n4.fcsf_labelname = 'DDH Lead Id'
left join FolderCustomerSpecificField_Table as n5 on n5.fcsf_projno = fm.projno
      and n5.fcsf_folderno = fm.folderno and n5.fcsf_labelname = 'DDH Member Number'
where fm.baid = '2' -- and fm.status in ('0','1','2')
  and convert(date, fm.folderdate) = format(convert(date,'{0}'),'MM/dd/yyyy') and atd.passname like '%{1}'"""

    cur_worker.execute(sel_sql.format(key_value[0], key_value[1]))
    row = cur_worker.fetchone()

    the_leadid = ''
    the_contract = ''
    results = ''

    while row:
        the_folder = row[0]
        the_order = row[1]
        # 4  = Lead Id
        # 5  = ContractNum
        # 6  = Source
        # 7  = LeadId4,
        # 8  = DDHMemberNum
        if row[4] is None:
            if row[7] is None or row[7] == '':
                the_leadid = ''
            else:
                the_leadid = row[7]
            if row[8] is None:
                the_leadid = ''
            else:
                the_leadid = row[8]
        else:
            the_leadid = row[4]

        if row[5]:
            the_contract = row[5]
        else:
            if row[6]:
                if 'US' in row[6]:
                    the_contract = row[6][2:]
                else:
                    the_contract = row[6]
        # if the_leadid
        results = [the_folder, the_order, the_leadid, the_contract]
        row = cur_worker.fetchone()
    return results


def getPassengerLastNameOnly(key_value, cur_worker):
    """Collect the ticket details based on the passenger last name"""

    sel_sql = """
    select atd.projno, atd.folderno, atd.passname, atd.ticketno,
       replace(replace(replace(replace(n1.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as LeadId,
       n2.fcsf_labeldata as ContractNum, n3.fcsf_labeldata as Source,
       replace(replace(replace(replace(n4.fcsf_labeldata,' ',''),'\t',''),'\n',''),'"','') as LeadId4,
       n5.fcsf_labeldata as DDHMemberNum
from AirTicketDetails as atd
inner join foldermaster as fm on fm.projno=atd.projno and fm.folderno = atd.folderno
left join FolderCustomerSpecificField_Table as n1 on n1.fcsf_projno = fm.projno
      and n1.fcsf_folderno = fm.folderno and n1.fcsf_labelname = 'Membership Number'
left join FolderCustomerSpecificField_Table as n2 on n2.fcsf_projno = fm.projno
      and n2.fcsf_folderno = fm.folderno and n2.fcsf_labelname = 'ContractNumber'
left join FolderCustomerSpecificField_Table as n3 on n3.fcsf_projno = fm.projno
      and n3.fcsf_folderno = fm.folderno and n3.fcsf_labelname = 'Source'
left join FolderCustomerSpecificField_Table as n4 on n4.fcsf_projno = fm.projno
      and n4.fcsf_folderno = fm.folderno and n4.fcsf_labelname = 'DDH Lead Id'
left join FolderCustomerSpecificField_Table as n5 on n5.fcsf_projno = fm.projno
      and n5.fcsf_folderno = fm.folderno and n5.fcsf_labelname = 'DDH Member Number'
where fm.baid = '2' and atd.passname like '%{0}'"""

    cur_worker.execute(sel_sql.format(key_value))
    row = cur_worker.fetchone()

    the_leadid = ''
    the_contract = ''
    results = ''

    while row:
        the_folder = row[0]
        the_order = row[1]
        # 4  = Lead Id
        # 5  = ContractNum
        # 6  = Source
        # 7  = LeadId4,
        # 8  = DDHMemberNum
        if row[4] is None:
            if row[7] is None or row[7] == '':
                the_leadid = ''
            else:
                the_leadid = row[7]
            if row[8] is None:
                the_leadid = ''
            else:
                the_leadid = row[8]
        else:
            the_leadid = row[4]

        if row[5]:
            the_contract = row[5]
        else:
            if row[6]:
                if 'US' in row[6]:
                    the_contract = row[6][2:]
                else:
                    the_contract = row[6]
        # if the_leadid
        results = [the_folder, the_order, the_leadid, the_contract]
        row = cur_worker.fetchone()
    return results


def getPassenger(key_values, cur_worker):
    """Check the PassengerMaster table for the transaction names"""

    sel_sql = """
select top 1 pm.projno, pm.folderno, pm.slno, pm.lastname, pm.firstname, fm.crdate, fm.folderdate, fm.inetref, fm.bookedby
from PassengerMaster as pm
inner join foldermaster as fm on fm.projno = pm.projno and fm.folderno = pm.folderno            
where fm.baid = '2' and pm.lastname = '{0}' and pm.firstname like '{1}%' order by pm.projno desc"""

    cur_worker.execute(sel_sql.format(key_values[0], key_values[1]))
    row = cur_worker.fetchone()

    if row is None:
        results = ''
    else:
        results = row
    return results


def getTicketValue(key_value):
    """Extract from the provided parameter the value of the AIR ticket which is length of 10"""

    return key_value[3:13]


def getticket(key_value, cur_worker):
    """Check if the Ticket Details from the AirTicketDetails are available based on the parameter provided"""

    sel_sql = """select atd.projno, atd.folderno, fm.inetref, atd.suppname, fm.bookedby, atd.passname, atd.ticketno, 
                 atd.payable, atd.vairlineairname, atd.selltot
from AirTicketDetails as atd
inner join foldermaster as fm on fm.projno=atd.projno and fm.folderno = atd.folderno
where fm.baid = '2' and atd.ticketno like '{0}'; """

    cur_worker.execute(sel_sql.format(key_value))
    row = cur_worker.fetchone()
    # 0 atd.projno
    # 1 atd.folderno
    # 2 fm.inetref
    # 3 atd.suppname
    # 4 atd.bookedby
    # 5 atd.passname
    # 6 atd.ticketno
    # 7 atd.payable
    # 8 atd.vairlineairname
    # 9 atd.selltot


    return row


def getPassengerName(trx_passenger):
    """Break the CC report passenger name and correct it for backend processing"""

    if trx_passenger and trx_passenger[-1] == 'M':
        trx_passenger = trx_passenger[:-1].strip()  # Remove the last letter as it is M and then remove extra spaces
    trx_passenger = trx_passenger.replace(' MRS', '')  # Remove the salutations from the end of the name string
    trx_passenger = trx_passenger.replace(' MR', '')
    trx_passenger = trx_passenger.replace(' MS', '')

    if '/' in trx_passenger:
        # Splite the name to last name and first name as delimited by a /
        results = trx_passenger.split('/')
    else:
        worker = trx_passenger.split(' ')
        if len(worker) == 3:
            results = [worker[0],worker[1] + ' ' + worker[2]]
        else:
            if len(worker) == 2:
                results = [worker[1], worker[0]]
            else:
                results = worker[0].strip()
        # Check if title part of name received and then remove the items
        # rebuild the results so that last name, first name, initial
        # handle the special case that the name is not delimited by a /
        # the_pax = interm_name.split('/')  # Split the name into a list so 0 is Last Name, 1 is First Name
    results = [itm.strip() for itm in results]
    return results


def geteVolveStuff(key_value, cur_worker):
    """Get the eVolve data record as we have a web reference"""

    # sel_sql = """select leadid, order_statusid, order_memid, order_createddatetime, prod_webref from wrk_hsv_orders where prod_webref = '{0}'"""
    sel_sql = """select leadid, n0.Field_Value as eLeadId, n1.Field_Value as eContract
from wrk_hsv_orders as od
left join wrk_hsv_OrderFields as n0 on n0.Field_OrderID = od.Order_ID and n0.Field_Name = 'LeadID'
left join wrk_hsv_OrderFields as n1 on n1.Field_OrderID = od.Order_ID and n1.Field_Name = 'ContractNumber'
    where prod_webref = '{0}'"""

    cur_worker.execute(sel_sql.format(key_value))
    row = cur_worker.fetchone()

    if row is None:
        results = None
    else:
        results = [row[0], row[1], row[2]]
    return results


def getPreferredPax_data(key_value, cur_worker):
    """Get the Pen Air data from the Other Details using the partial PAX name, ([last_name, row[3]], cursor_worker)"""

    sel_sql = """select projno, folderno, description, rate, suppname from othersdetails
where projno in (select projno from passengermaster where lastname like '{0}%')
and rate = {1};
    """

    cur_worker.execute(sel_sql.format(key_value[0], key_value[1]))
    row = cur_worker.fetchone()

    if row is None:
        results = None
    else:
        results = [row[0], row[1], row[2], row[3], row[4]]
    return results


def generate_report(runstyle, conn_main):
    """Generate a workbook for review by operations"""

    cursor_main = conn_main.cursor()

    sel_sql = """select Transaction_ID, Record_Type, Merchant, Merchant_Location, Account_Number, MCC, Date_Occurred, 
    Date_Posted, Original_Amount, Original_Currency_Code, Conversion_Rate, Settlement_Amount, Allocation, Transaction_Description, 
    Reference_Number, Purch_Description, Purch_Quantity, Purch_Unit_Cost, Purch_Unit_Of_Measure, Purch_Extended_Amount, 
    Passenger_Name, Ticket_Number, Travel_Date, Travel_Legs, import_date, projno, folderno, leadid, contract_number, 
    product_type,webref, booking_agent, supplier from wrk_hsv_cc_transactions; """

    # Create New Report file
    report_file = datetime.now().strftime('%Y-%m-%d') + '_HSV_reconciliation_' + runstyle + '.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    buildheadings( worksheet )

    cursor_main.execute(sel_sql)
    main_row = cursor_main.fetchone()  # Collect the initial row to start the processing

    # Ok now load the data onto the report workbook
    itm = 1
    while main_row:
        itm += 1
        col = 0
        for dat in main_row:
            col += 1
            worksheet.cell(row=itm, column=col, value=dat)

        main_row = cursor_main.fetchone()

    workbook.save(report_file)

    # So report generated, now email the thing
    if platform.release() == '10':
        msg_to = 'ajay@synchroworks.net'
        msg_cc = msg_to
        #msg_subject = ' TESTING: HSV CC Reconciliation Report for %s' % (runstyle)
    else:
        msg_to = email_to
        msg_cc = email_cc

    msg_subject = 'HSV CC Reconciliation Report for %s' % (runstyle)

    msg_text = 'The attached workbook contains the HSV CC Transactions for the current period.'

    msg_html = f'<html><head></head><body><p>Good Day<br><br><bold>{msg_text}</bold></body></html>'

    serefin.send_mail(email_from, msg_to, msg_cc, msg_subject, msg_text, msg_html, report_file, email_server,
                      587, None, email_user, email_pass, True, True
                      )


def process_paxticket_data(runstyle, conn_main, conn_writer, conn_secondary, conn_worker):
    """Process the wrk_HSV_TRANSACTIONS table to find the ProjNo/FolderNo combinations"""

    # Get all available records from work table
    cursor_main = conn_main.cursor()  # MySQL connection to wrkhsv_cc_transactions
    cursor_secondary = conn_secondary.cursor()  # MySQL connection as a worker thread

    cursor_worker = conn_worker.cursor()  # MS SQL connection to the PenAir database

    cursor_writer = conn_writer.cursor()

    # row[0] TRX Id
    # row[1] Name on Transaction
    # row[2] Ticket Number
    # row[3] Booking Date
    # row[4] Travel Date
    # row[5] trx amount

    # Pull specific fields from teh CC stement provided by HSV, twice weekly
    sel_sql = """SELECT transaction_id, passenger_name, ticket_number, date_occurred, travel_date, settlement_amount, 
    Merchant 
    FROM wrk_hsv_cc_transactions where (leadid = '' or leadid is null) and (projno = '' or projno is null)
    and (len(Passenger_Name) > 1 and len(Ticket_Number) > 1) 
    and Passenger_name not like '%/PREFERRED ZONE'; """

    cursor_main.execute( sel_sql )
    main_row = cursor_main.fetchone()  # Collect the initial row to start the processing

    while main_row:
        # Initialize key attributes, may not be required but hey be safe
        product_type = ''
        webref = ''
        the_agent = ''
        the_folderno = ''
        the_orderno = ''
        the_supplier = ''
        the_lead = ''
        the_contract = ''
        fm_details = None
        the_pax = None
        yourref = None

        trx_id = main_row[0]
        ticket_number = getTicketValue( main_row[2] )
        travel_date = main_row[4]

        if main_row[1] != '':
            the_pax = getPassengerName( main_row[1].upper() )

        if main_row[1] == '' or main_row[2] == '':
            # Handle the missing ticket and PAX using the Merchant details

            # merchant_name = re.findall('[a-zA-Z]+', main_row[6] )  # disable code as merchant_name not used
            merchant_num = re.findall('[0-9]+', main_row[6])
            try:
                if main_row[4].isdigit():
                    yourref = merchant_num[0]
            except Exception as e:
                    yourref = None

            if yourref:
                product_type = 'HOTEL'
                fm_details = getFolderMasterDetails(merchant_num[0], cursor_worker)

            if fm_details:
                the_folderno = fm_details[0]
                the_orderno = fm_details[1]
                webref = fm_details[2]
                pax_details = fm_details[3]
                the_pax = pax_details.split()
                the_agent = fm_details[4]
                travel_date = fm_details[5]
                ticket_number = None

        if ticket_number:
            ticket_details = getticket(ticket_number, cursor_worker)
            if ticket_details:
                product_type = 'AIR'
                the_folderno = ticket_details[0]
                the_orderno = ticket_details[1]
                webref = ticket_details[2]
                the_supplier = ticket_details[3]
                the_agent = ticket_details[4]
            else:
                # Find the information based on the PAX name
                # 0 pm.projno,
                # 1 pm.folderno,
                # 2 pm.slno,
                # 3 pm.lastname,
                # 4 pm.firstname,
                # 5 fm.crdate,
                # 6 fm.folderdate,
                # 7 fm.inetref,
                # 8 fm.bookedby
                pax_details = getPassenger(the_pax, cursor_worker)
                if pax_details:
                    the_folderno = pax_details[0]
                    the_orderno = pax_details[1]
                    webref = pax_details[7]
                    the_agent = pax_details[8]

        if the_folderno:
            the_member = getLeadMemberShipId([the_folderno, the_orderno], cursor_worker)
            if the_member:
                the_lead = the_member[0]
                the_contract = the_member[1]

        if the_supplier == '':
            other_details = getOtherdetails([the_folderno, the_orderno], cursor_worker)
            if other_details:
                the_supplier = other_details[3]

        if (the_lead == None or the_lead == '') and webref != '':
            the_evolve_details = geteVolveStuff(f"ACC-{webref}", cursor_secondary)
            if the_evolve_details:
                the_lead = the_evolve_details[1]
                the_contract = the_evolve_details[2]

        if the_lead == 'None' or the_lead == None:
            the_lead=''
        if the_lead == '' and the_pax != None:
            results = getTravelDate([travel_date,the_pax[0]], cursor_worker)
            if results:
                if the_folderno == '':
                    the_folderno = results[0]
                if the_orderno == '':
                    the_orderno = results[1]
                if the_lead == '':
                    the_lead = results[2]
                if the_contract == '':
                    the_contract = results[3]
            else:
                if the_pax != None:
                    results = getPassengerLastNameOnly(the_pax[0], cursor_worker)
                    if results:
                        if the_folderno == '':
                            the_folderno = results[0]
                        if the_orderno == '':
                            the_orderno = results[1]
                        if the_lead == '':
                            the_lead = results[2]
                        if the_contract == '':
                            the_contract = results[3]

        if the_lead != '' and the_contract == '':
            the_contract = getContractNumber(the_lead, cursor_secondary)

        if the_contract != '' and the_lead == '':
            the_lead = getLeadid(the_contract, cursor_secondary)

        #
        # do the update
        if len(the_lead) > 2:
            upd_sql = """UPDATE wrk_hsv_cc_transactions set projno='{1}', folderno ='{2}', leadid='{3}', 
                                contract_number='{4}', product_type='{5}', webref='{6}', booking_agent='{7}', supplier='{8}'
                                where transaction_id='{0}'"""
            try:
                cursor_writer.execute(
                    upd_sql.format(trx_id, the_folderno, the_orderno, the_lead, the_contract, product_type, webref, the_agent, the_supplier))
            except Exception as e:
                print(f"Details : {trx_id}")
                print_exception()

            conn_writer.commit()

        main_row = cursor_main.fetchone()


def process_merchant_data(runstyle, conn_main, conn_writer, conn_secondary, conn_worker):
    """Process the wrk_HSV_TRANSACTIONS table to find the ProjNo/FolderNo combinations"""

    # Get all available records from work table
    cursor_main = conn_main.cursor()
    cursor_secondary = conn_secondary.cursor()
    cursor_worker = conn_worker.cursor()

    cursor_writer = conn_writer.cursor()

    # Pull specific fields from teh CC stement provided by HSV, twice weekly
    sel_sql = """SELECT transaction_id, date_occurred, travel_date, settlement_amount, Merchant 
    FROM wrk_hsv_cc_transactions 
    where (leadid = '' or leadid is null) and (projno = '' or projno is null) and (len(Passenger_Name) < 2 and len(Ticket_Number) < 2) 
    and Passenger_name not like '%/PREFERRED ZONE' ; 
    """

    cursor_main.execute( sel_sql )
    main_row = cursor_main.fetchone()  # Collect the initial row to start the processing

    while main_row:
        # Initialize key attributes, may not be required but hey be safe
        product_type = ''
        webref = ''
        the_agent = ''
        the_folderno = ''
        the_orderno = ''
        the_supplier = ''
        the_lead = ''
        the_contract = ''
        fm_details = None
        the_pax = None
        yourref = None

        # Handle the missing ticket and PAX using the Merchant details
        trx_id = main_row[0]
        # merchant_details = main_row[4].split(' ')
        # yourref = merchant_details[1]
        # merchant_name = re.findall('[a-zA-Z]+', main_row[4])  ## disabled as not used any other place
        merchant_num = re.findall('[0-9]+', main_row[4])
        # if main_row[4].isdigit():
        try:
            if merchant_num[0].isdigit():
                yourref = merchant_num[0]
        except Exception as e:
            yourref = None

        if yourref:
            product_type = 'HOTEL'
            fm_details = getFolderMasterDetails(merchant_num[0], cursor_worker)

            # Laygur for FolderMaster
            # fm.projno, fm.folderno, fm.inetref, fm.pass, fm.bookedby, fm.folderdate
            #      0            1           2        3           4             5

        if fm_details:
            the_folderno = fm_details[0]
            the_orderno = fm_details[1]
            webref = fm_details[2]
            pax_details = fm_details[3]
            the_pax = pax_details.split()
            the_agent = fm_details[4]
            travel_date = fm_details[5]
            ticket_number = None
            #
            # Get the Lead and Contract
            results = getLeadContractDetails([the_folderno, the_orderno], cursor_worker)
            if results:
                the_lead = results[0]
                the_contract = results[1]

        if the_lead == '' or the_contract == '':
            if the_folderno != '' or the_orderno != '':
                the_member = getLeadMemberShipId([the_folderno, the_orderno], cursor_worker)
                if the_member:
                    if the_lead == '':
                        the_lead = the_member[0]
                    if the_contract == '':
                        the_contract = the_member[1]

        if the_supplier == '':
            other_details = getOtherdetails([the_folderno, the_orderno], cursor_worker)
            if other_details:
                the_supplier = other_details[3]

        if (the_lead == None or the_lead == '') and webref != '':
            the_evolve_details = geteVolveStuff(f"ACC-{webref}", cursor_secondary)
            if the_evolve_details:
                the_lead = the_evolve_details[1]
                the_contract = the_evolve_details[2]

        if the_lead == 'None' or the_lead is None:
            the_lead = ''
        if the_lead == '' and the_pax != None:
            results = getTravelDate([travel_date, the_pax[0]], cursor_worker)
            if results:
                if the_folderno == '':
                    the_folderno = results[0]
                if the_orderno == '':
                    the_orderno = results[1]
                if the_lead == '':
                    the_lead = results[2]
                if the_contract == '':
                    the_contract = results[3]
            else:
                if the_pax != None:
                    results = getPassengerLastNameOnly(the_pax[0], cursor_worker)
                    if results:
                        if the_folderno == '':
                            the_folderno = results[0]
                        if the_orderno == '':
                            the_orderno = results[1]
                        if the_lead == '':
                            the_lead = results[2]
                        if the_contract == '':
                            the_contract = results[3]

        if the_lead != '' and the_contract == '':
            the_contract = getContractNumber(the_lead, cursor_secondary)

        if the_contract != '' and the_lead == '':
            the_lead = getLeadid(the_contract, cursor_secondary)

        #
        # do the update
        if len(the_lead) > 2:
            upd_sql = """UPDATE wrk_hsv_cc_transactions set projno='{1}', folderno ='{2}', leadid='{3}', 
                                contract_number='{4}', product_type='{5}', webref='{6}', booking_agent='{7}', supplier='{8}'
                                where transaction_id='{0}'"""
            try:
                cursor_writer.execute(
                    upd_sql.format(trx_id, the_folderno, the_orderno, the_lead, the_contract, product_type, webref, the_agent, the_supplier))
            except Exception as e:
                print(f"Details : {trx_id}")
                print_exception()

            conn_writer.commit()

        main_row = cursor_main.fetchone()


def process_preferred_zone_data(runstyle, conn_main, conn_writer, conn_secondary, conn_worker):
    """Process the wrk_HSV_TRANSACTIONS table to find the ProjNo/FolderNo combinations"""

    # Get all available records from work table
    cursor_main = conn_main.cursor()  # MySQL connection to wrkhsv_cc_transactions
    cursor_secondary = conn_secondary.cursor()  # MySQL connection as a worker thread

    cursor_worker = conn_worker.cursor()  # MS SQL connection to the PenAir database

    cursor_writer = conn_writer.cursor()

    # Pull specific fields from the CC statement provided by HSV, twice weekly
    sel_sql = """SELECT transaction_id, passenger_name, travel_date, settlement_amount,
    Merchant FROM wrk_hsv_cc_transactions
    where (leadid = '' or leadid is null) and (projno = '' or projno is null) and Passenger_name like '%/PREFERRED ZONE';
    """
    # sel_sql = """SELECT transaction_id, passenger_name, travel_date, settlement_amount,
    # Merchant FROM wrk_hsv_cc_transactions
    # where transaction_id ='857929344' ;
    # """
    cursor_main.execute( sel_sql )
    main_row = cursor_main.fetchone()  # Collect the initial row to start the processing

    while main_row:
        # Initialize key attributes, may not be required but hey be safe
        the_folderno = ''
        the_orderno = ''
        the_lead = ''
        the_contract = ''
        yourref = None
        folder_data = None

        passenger_name = main_row[1].split('/')
        last_name = passenger_name[0].strip()
        trx_id = main_row[0]
        zone_data = getPreferredPax_data([last_name, main_row[3]], cursor_worker)

        if zone_data:
            the_folderno = zone_data[0]
            the_orderno = zone_data[1]
            product_type = zone_data[2]
            the_supplier = zone_data[4]

            # fm.projno, fm.folderno, fm.inetref, fm.pass, fm.bookedby, fm.folderdate
            folder_data = getFolderMasterbyProjno(the_folderno, cursor_worker)
            the_agent = folder_data[4]
            yourref = folder_data[2]

            if the_folderno:
                the_member = getLeadMemberShipId([the_folderno, the_orderno], cursor_worker)
                if the_member:
                    the_lead = the_member[0]
                    the_contract = the_member[1]


            if the_lead != '' and the_contract == '':
                the_contract = getContractNumber(the_lead, cursor_secondary)

            if the_contract != '' and the_lead == '':
                the_lead = getLeadid(the_contract, cursor_secondary)

            #
            # do the update
            if len(the_lead) > 2:
                upd_sql = """UPDATE wrk_hsv_cc_transactions set projno='{1}', folderno ='{2}', leadid='{3}', 
                                    contract_number='{4}', product_type='{5}',webref='{6}', booking_agent='{7}', supplier='{8}'
                                    where transaction_id='{0}'"""
                try:
                    cursor_writer.execute(
                        upd_sql.format(trx_id, the_folderno, the_orderno, the_lead, the_contract, product_type, yourref, the_agent, the_supplier))
                except Exception as e:
                    print(f"Details : {trx_id}")
                    print_exception()

                conn_writer.commit()

        main_row = cursor_main.fetchone()


def openMSconnection(drv, server, db, user, pwd):
    """Open database connection to Microsoft SQL Server"""

    try:
        conn = pyodbc.connect(f"DRIVER={drv};SERVER={server};PORT=1433;DATABASE={db};UID={user};PWD={pwd};")
    except Exception as e:
        print_exception()
        sys.exit(1)

    return conn


def main(argv):

    runstyle = 'OCTOBER'
    include_report = True

    # try:
        # opts, args = getopt.getopt(argv,"", ["run=", "report"])
    # except getopt.GetoptError:
        # print('match_cc_trx_backend.py --run <type_of_run> [--report]')
        # print_exception()
        # sys.exit(2)

    # for opt, arg in opts:
        # if opt == '-h':
            # print('match_cc_trx_backend.py --run <type_of_run>  [--report]')
            # sys.exit()
        # elif opt == '--report':
            # include_report = True
        # elif opt == '--run':
            # runstyle = arg

    conn_main = openMSconnection(ms_db_drv, ms_svr, ms_db, ms_usr, ms_pwd)  # NA instance access
    conn_secondary = openMSconnection(ms_db_drv, ms_svr, ms_db, ms_usr, ms_pwd)  # NA instance access
    conn_worker = openMSconnection(msp_db_drv, msp_svr, msp_db_na, msp_usr, msp_pwd)  # NA instance access
    conn_writer = openMSconnection(ms_db_drv, ms_svr, ms_db, ms_usr, ms_pwd)  # NA instance access

    # Now process the data
    process_paxticket_data(runstyle, conn_main, conn_writer, conn_secondary, conn_worker)
    process_merchant_data(runstyle, conn_main, conn_writer, conn_secondary, conn_worker)
    process_preferred_zone_data(runstyle, conn_main, conn_writer, conn_secondary, conn_worker)

    conn_main.close()
    conn_secondary.close()
    conn_worker.close()
    conn_writer.close()

    if include_report:
        conn_main = openMSconnection(ms_db_drv, ms_svr, ms_db, ms_usr, ms_pwd)  # NA instance access
        generate_report(runstyle, conn_main)


if __name__ == "__main__":
    main(sys.argv[1:])
